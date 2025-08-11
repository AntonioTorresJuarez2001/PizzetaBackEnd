# productos/views.py
import csv
import io
import unicodedata
from typing import Dict, Any, List

from django.db import transaction, connection
from django.db.models import ProtectedError

from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi

from productos.models import Producto
from .serializers import ProductoSerializer

from usuarios.permissions import EmpleadoSoloLecturaPermission
from usuarios.utils.roles import check_dueno


# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

def _norm(s: str) -> str:
    """Normaliza texto: trim, quita acentos, minúsculas."""
    if s is None:
        return ""
    s = " ".join(str(s).strip().split())
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return s.lower()


def _reset_pk_sequence_portable(model):
    """
    Ajusta el contador de autoincremento después de insertar IDs manuales.
    - MySQL: ALTER TABLE ... AUTO_INCREMENT = MAX(id)+1 (solo sube)
    - Postgres: setval(...)
    - Otros: no-op
    """
    table = model._meta.db_table
    pk_col = model._meta.pk.column
    vendor = connection.vendor

    with connection.cursor() as cursor:
        cursor.execute(f"SELECT COALESCE(MAX({pk_col}), 0) + 1 FROM {table}")
        next_id = cursor.fetchone()[0] or 1

        if vendor == "mysql":
            cursor.execute(
                "SELECT AUTO_INCREMENT "
                "FROM information_schema.tables "
                "WHERE table_schema = DATABASE() AND table_name = %s",
                [table],
            )
            row = cursor.fetchone()
            current_ai = (row[0] if row else None) or 1
            if next_id > current_ai:
                cursor.execute(f"ALTER TABLE {table} AUTO_INCREMENT = %s", [int(next_id)])

        elif vendor == "postgresql":
            cursor.execute(
                "SELECT setval(pg_get_serial_sequence(%s, %s), %s, true);",
                [table, pk_col, int(next_id) - 1 if next_id > 1 else 1],
            )
        else:
            pass  # SQLite y otros: sin acción


# ─────────────────────────────────────────────────────────────
# CRUD de Productos (anidados por pizzería)
# ─────────────────────────────────────────────────────────────

class ProductoListCreateByPizzeriaAPIView(generics.ListCreateAPIView):
    """
    Vista API para listar y crear productos asociados a una pizzería específica.

    Esta vista combina dos funcionalidades:
    - **GET**: Lista todos los productos de una pizzería específica.
    - **POST**: Crea un nuevo producto en la pizzería indicada (solo el dueño puede hacerlo).

    Reglas de acceso:
    - Solo usuarios autenticados pueden acceder.
    - Empleados tienen permiso de solo lectura (GET).
    - Solo el dueño de la pizzería puede crear productos (POST).

    Parámetros de URL:
    - pizzeria_id (int): Identificador único de la pizzería.

    Respuestas:
    - 200 OK: Lista de productos (GET).
    - 201 Created: Producto creado exitosamente (POST).
    - 403 Forbidden: Si el usuario no tiene permisos.
    - 404 Not Found: Si la pizzería no existe.
    """

    permission_classes = [IsAuthenticated, EmpleadoSoloLecturaPermission]
    serializer_class = ProductoSerializer

    @swagger_auto_schema(tags=["Productos"])
    def get(self, request, *args, **kwargs):
        """
        Obtiene la lista de todos productos de la pizzería especificada.

        URL de ejemplo:
            GET /pizzerias/{pizzeria_id}/productos/

        Respuesta:
        [
            {
                "id": 1,
                "id_externo": 101,
                "nombre": "Pizza Hawaiana",
                "precio": "150.00",
                "categoria": "Pizza",
                "descripcion": "Pizza con piña y jamón",
                "activo": true
            },
            ...
        ]
        """
        return super().get(request, *args, **kwargs)

    @swagger_auto_schema(tags=["Productos"])
    def post(self, request, *args, **kwargs):
        """
        Crea un nuevo producto en la pizzería especificada.

        URL de ejemplo:
            POST http://localhost:8001/api/pizzerias/{pizzeria_id}/productos/

        Body de ejemplo:
        {
            "id_externo": 101, <-- id_pro de Martin
            "nombre": "Pizza Hawaiana",
            "precio": "150.00",
            "categoria": "Pizza",
            "descripcion": "Pizza con piña y jamón",
            "activo": true
        }

        Respuesta:
        {
            "id": 5,
            "id_externo": 101,
            "nombre": "Pizza Hawaiana",
            "precio": "150.00",
            "categoria": "Pizza",
            "descripcion": "Pizza con piña y jamón",
            "activo": true
        }
        """
        return super().post(request, *args, **kwargs)

    def get_queryset(self):
        """
        Retorna el queryset de productos filtrado por `pizzeria_id`.
        - Si `pizzeria_id` no está presente en la URL, retorna queryset vacío.

        Retorna:
            QuerySet de objetos Producto asociados a la pizzería indicada.
        """
        if getattr(self, "swagger_fake_view", False):
            return Producto.objects.none()

        pizzeria_id = self.kwargs.get("pizzeria_id")
        if not pizzeria_id:
            return Producto.objects.none()

        # Lectura: permitido por EmpleadoSoloLecturaPermission.
        # Escritura: se valida en perform_create.
        return Producto.objects.filter(pizzeria_id=pizzeria_id)

    def perform_create(self, serializer):
        """
        Lógica personalizada para crear un producto.

        Pasos:
        1. Obtiene el `pizzeria_id` de la URL.
        2. Verifica que el usuario autenticado sea dueño de la pizzería 
           (`check_dueno` lanza excepción si no lo es).
        3. Guarda el nuevo producto asignándolo a la pizzería correspondiente.

        Parámetros:
            serializer (ProductoSerializer): Instancia del serializador validada.
        """
        pizzeria_id = self.kwargs["pizzeria_id"]
        # Solo dueño puede crear
        check_dueno(self.request.user, pizzeria_id)
        serializer.save(pizzeria_id=pizzeria_id)



class ProductoRetrieveUpdateDestroyByPizzeriaAPIView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated, EmpleadoSoloLecturaPermission]
    serializer_class = ProductoSerializer
    lookup_url_kwarg = "pk"

    # Swagger: etiqueta las operaciones
    @swagger_auto_schema(tags=["Productos"])
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    @swagger_auto_schema(tags=["Productos"])
    def put(self, request, *args, **kwargs):
        return super().put(request, *args, **kwargs)

    @swagger_auto_schema(tags=["Productos"])
    def patch(self, request, *args, **kwargs):
        return super().patch(request, *args, **kwargs)

    @swagger_auto_schema(tags=["Productos"])
    def delete(self, request, *args, **kwargs):
        return super().delete(request, *args, **kwargs)

    def get_queryset(self):
        # 1) Cuando drf-yasg genera el esquema, no hay kwargs -> corta y evita side-effects
        if getattr(self, "swagger_fake_view", False):
            return Producto.objects.none()

        # 2) Evita KeyError si por alguna razón no viene la ruta completa
        pizzeria_id = self.kwargs.get("pizzeria_id")
        if not pizzeria_id:
            return Producto.objects.none()

        # Lectura: permitido por EmpleadoSoloLecturaPermission.
        # Escrituras: se validan en perform_update/perform_destroy.
        return Producto.objects.filter(pizzeria_id=pizzeria_id)

    def perform_update(self, serializer):
        # Solo dueño puede actualizar
        pizzeria_id = self.kwargs["pizzeria_id"]
        check_dueno(self.request.user, pizzeria_id)
        serializer.save()

    def perform_destroy(self, instance):
        # Solo dueño puede borrar
        check_dueno(self.request.user, instance.pizzeria_id)
        try:
            instance.delete()
        except ProtectedError:
            # Responder 409 Conflict cuando hay relaciones protegidas
            from rest_framework.exceptions import APIException
            class Conflict(APIException):
                status_code = 409
                default_detail = "No puedes eliminar un producto que ya ha sido vendido."
            raise Conflict()


# ─────────────────────────────────────────────────────────────
# Importación masiva (CSV/Excel)
# ─────────────────────────────────────────────────────────────

class ImportProductosAPIView(APIView):
    """
    Importa productos por pizzería desde CSV/Excel.

    Reglas:
      - 'Id_Pro' del archivo se usa como PK interno (Producto.id).
      - 'Nombre' -> Producto.nombre
      - 'Línea'  -> Producto.categoria

    Si existe ese Id_Pro en esa pizzería, se actualiza nombre/categoría (no toca precio/activo).
    Si no existe, se crea con precio=0 y activo=True.
    """
    permission_classes = [IsAuthenticated, EmpleadoSoloLecturaPermission]
    parser_classes = [MultiPartParser, FormParser]

    @swagger_auto_schema(
        tags=["Productos"],
        operation_summary="Importar productos por pizzería (CSV o Excel)",
        consumes=["multipart/form-data"],
        manual_parameters=[
            openapi.Parameter(
                name="pizzeria_id",
                in_=openapi.IN_PATH,
                type=openapi.TYPE_INTEGER,
                required=True,
                description="ID de la pizzería destino",
            ),
            openapi.Parameter(
                name="archivo",
                in_=openapi.IN_FORM,
                type=openapi.TYPE_FILE,
                required=True,
                description="Archivo .csv, .xlsx o .xls con columnas: Id_Pro, Nombre, Línea",
            ),
        ],
        responses={
            200: "OK",
            207: "Multi-Status (algunas filas con error)",
            400: "Bad request",
            403: "Forbidden",
        },
    )
    def post(self, request, *args, **kwargs):
        pizzeria_id = kwargs.get("pizzeria_id")
        # Solo dueño puede importar (escritura)
        check_dueno(request.user, pizzeria_id)

        up = request.FILES.get("archivo")
        if not up:
            return Response({"detail": "Falta 'archivo'."}, status=status.HTTP_400_BAD_REQUEST)

        name = (up.name or "").lower()
        try:
            if name.endswith(".csv"):
                rows = self._read_csv(up)
            elif name.endswith(".xlsx") or name.endswith(".xls"):
                rows = self._read_excel(up)  # robusto: detecta encabezado real
            else:
                return Response(
                    {"detail": "Formato no soportado. Usa .csv o .xlsx/.xls."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except Exception as e:
            return Response({"detail": f"Error leyendo archivo: {e}"}, status=status.HTTP_400_BAD_REQUEST)

        header_map = self._resolve_header_map(rows["headers"])
        missing = [k for k, v in header_map.items() if v is None]
        if missing:
            return Response(
                {"detail": f"Columnas faltantes: {', '.join(missing)}. Requeridas: Id_Pro, Nombre, Línea."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created = 0
        updated = 0
        errors: List[Dict[str, Any]] = []
        sample_out = []
        seen_ids = set()  # detecta duplicados dentro del archivo

        with transaction.atomic():
            for idx, r in enumerate(rows["data"], start=2):  # asumiendo encabezado en fila 1
                try:
                    id_pro_raw = r.get(header_map["Id_Pro"])
                    nombre_raw = r.get(header_map["Nombre"])
                    linea_raw = r.get(header_map["Línea"])

                    if id_pro_raw in (None, ""):
                        raise ValueError("Id_Pro vacío.")
                    try:
                        id_pro = int(str(id_pro_raw).strip())
                    except Exception:
                        raise ValueError("Id_Pro debe ser numérico entero.")

                    if id_pro <= 0:
                        raise ValueError("Id_Pro debe ser un entero positivo.")

                    if id_pro in seen_ids:
                        raise ValueError(f"Id_Pro repetido en el archivo: {id_pro}.")
                    seen_ids.add(id_pro)

                    nombre = str(nombre_raw or "").strip()
                    if not nombre:
                        raise ValueError("Nombre vacío.")

                    categoria = str(linea_raw or "").strip()
                    if not categoria:
                        raise ValueError("Línea (categoría) vacía.")

                    obj, exists = self._create_or_update_producto(
                        pizzeria_id=pizzeria_id, id_pro=id_pro, nombre=nombre, categoria=categoria
                    )
                    if exists:
                        updated += 1
                    else:
                        created += 1

                    if len(sample_out) < 5:
                        sample_out.append({"id": obj.id, "nombre": obj.nombre, "categoria": obj.categoria})

                except Exception as e:
                    errors.append({"row": idx, "error": str(e)})

            # Ajusta AUTO_INCREMENT/sequence según el motor (MySQL/Postgres)
            _reset_pk_sequence_portable(Producto)

        status_code = status.HTTP_200_OK if not errors else status.HTTP_207_MULTI_STATUS
        return Response(
            {"created": created, "updated": updated, "errors": errors, "sample": sample_out},
            status=status_code,
        )

    # ── Lectores ──────────────────────────────────────────────

    def _read_csv(self, uploaded_file) -> Dict[str, Any]:
        raw = uploaded_file.read()
        text = None
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text = raw.decode(enc)
                break
            except Exception:
                continue
        if text is None:
            raise ValueError("No se pudo decodificar el CSV.")

        # Detecta delimitador
        try:
            dialect = csv.Sniffer().sniff(text.splitlines()[0])
        except Exception:
            dialect = csv.excel
        reader = csv.reader(io.StringIO(text), dialect)

        rows = list(reader)
        if not rows:
            return {"headers": [], "data": []}

        headers = rows[0]
        data = [dict(zip(headers, row)) for row in rows[1:]]
        return {"headers": headers, "data": data}

    def _read_excel(self, uploaded_file) -> Dict[str, Any]:
        # Detección robusta de encabezado (por si vienen “Unnamed”)
        try:
            import openpyxl
        except ImportError:
            raise ImportError("Falta dependencia 'openpyxl'. Instala con: pip install openpyxl")

        wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        ws = wb.active

        peek = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            peek.append(list(row))
            if i >= 20:
                break
        if not peek:
            return {"headers": [], "data": []}

        def norm(s):
            if s is None:
                return ""
            s = " ".join(str(s).strip().split())
            s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
            return s.lower()

        header_row_idx = None
        candidates = ("id", "id pro", "id_pro", "idpro", "nombre", "linea", "línea", "categoria", "categoría", "producto")
        for idx, row in enumerate(peek):
            normalized = [norm(c) for c in row]
            score = sum(any(c in col for c in candidates) for col in normalized)
            if score >= 2:
                header_row_idx = idx
                break
        if header_row_idx is None:
            header_row_idx = 0

        all_rows = list(ws.iter_rows(values_only=True))
        headers = [str(h) if h is not None else "" for h in all_rows[header_row_idx]]
        data_rows = all_rows[header_row_idx + 1:]

        data = []
        for r in data_rows:
            d = {}
            for i, h in enumerate(headers):
                d[h] = r[i] if (r and i < len(r)) else None
            data.append(d)
        return {"headers": headers, "data": data}

    def _resolve_header_map(self, headers: List[str]) -> Dict[str, str]:
        norm_headers = {_norm(h): h for h in headers}
        def find(*cands):
            for c in cands:
                if c in norm_headers:
                    return norm_headers[c]
            return None
        return {
            "Id_Pro": find("id_pro", "id pro", "idpro", "id", "id producto", "id_producto"),
            "Nombre": find("nombre", "producto", "nombre producto", "nombre_producto", "descripcion", "descripción"),
            "Línea": find("linea", "línea", "categoria", "categoría", "familia", "rubro"),
        }

    # ── Upsert ────────────────────────────────────────────────
    def _create_or_update_producto(self, pizzeria_id: int, id_pro: int, nombre: str, categoria: str, strategy: str = "skip"):
        # Buscar por id_externo dentro de la pizzería
        try:
            obj = Producto.objects.get(pizzeria_id=pizzeria_id, id_externo=id_pro)
            changed = False
            if obj.nombre != nombre:
                obj.nombre = nombre; changed = True
            if obj.categoria != categoria:
                obj.categoria = categoria; changed = True
            if changed:
                obj.save(update_fields=["nombre", "categoria"])
            return obj, True
        except Producto.DoesNotExist:
            pass

        # Conflicto por nombre en la pizzería
        dup_qs = Producto.objects.filter(pizzeria_id=pizzeria_id, nombre__iexact=nombre)
        if dup_qs.exists():
            if strategy == "rename":
                nombre = self._next_available_name(pizzeria_id, nombre)
            elif strategy == "update_by_name":
                obj = dup_qs.first()
                changed = False
                if obj.categoria != categoria:
                    obj.categoria = categoria; changed = True
                if obj.nombre != nombre:
                    obj.nombre = nombre; changed = True
                if changed:
                    obj.save(update_fields=["nombre", "categoria"])
                return obj, True
            else:
                raise ValueError("Conflicto: ya existe un producto con ese nombre en esta pizzería.")

        # Crear nuevo (sin forzar PK)
        obj = Producto(
            pizzeria_id=pizzeria_id,
            id_externo=id_pro,     # <- Id_Pro ahora vive aquí
            nombre=nombre,
            categoria=categoria,
            precio=0,
            activo=True,
        )
        obj.save()
        return obj, False

